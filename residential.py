#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation


def make(out: Union[str, Path] = "swale_calculator.xlsx") -> Path:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True  # force Excel recalculation on open

    normal_style = wb._named_styles[0]
    normal_font = copy(normal_style.font)
    normal_font.name = "Roboto"
    normal_font.size = 10
    normal_style.font = normal_font

    ws = wb.active
    ws.title = "Swale Calculator"
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.25, footer=0.25)

    # -----------------------------
    # Styles (simple, inline)
    # -----------------------------
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Use opaque ARGB fills (FF......) so Excel shows them
    fill_header = PatternFill("solid", fgColor=Color(theme=9, tint=0.8))
    fill_input = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
    fill_deprecated = PatternFill(fill_type="solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
    deprecate_font = Font(color="FF7F7F7F", strike=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(widths: dict[str, float]) -> None:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def apply_border(cell_range: str) -> None:
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border

    # -----------------------------
    # Layout
    # -----------------------------
    set_col_width(
        {
            "A": 20,
            "B": 11,
            "C": 10,
            "D": 10,
            "E": 10,
            "F": 11,
            "G": 11,
            "H": 12,
            "I": 11,
        }
    )

    # -----------------------------
    # SITE STATISTICS
    # -----------------------------
    ws["A1"] = "SITE STATISTICS"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:C1")

    ws["A2"] = "Description"
    ws["B2"] = "Sq. Feet"
    ws["C2"] = "Acres"
    ws["D2"] = "Percent"
    for addr in ("A2", "B2", "C2", "D2"):
        ws[addr].font = header_font
        ws[addr].alignment = center
        ws[addr].fill = fill_header

    site_rows = [
        ("Lot size", 3),
        ("Building (incl. patios)", 4),
        ("Driveway", 5),
        ("Equipment pads", 6),
        ("Other impervious", 7),
    ]

    for label, r in site_rows:
        ws[f"A{r}"] = label
        ws[f"A{r}"].alignment = left

        ws[f"B{r}"].number_format = "#,##0"
        ws[f"B{r}"].alignment = right

        ws[f"C{r}"] = f'=IF(OR(B{r}="",B{r}=0),"",B{r}/43560)'
        ws[f"C{r}"].number_format = "0.000"
        ws[f"C{r}"].alignment = right

        # Percent: blank for Lot size row, else fraction of lot
        ws[f"D{r}"] = "" if r == 3 else f'=IF(OR(B{r}="",B{r}=0,$B$3="",$B$3=0),"",B{r}/$B$3)'
        ws[f"D{r}"].number_format = "0.0%"
        ws[f"D{r}"].alignment = right

    # Seed example values (edit as needed)
    ws["B3"] = 6741
    ws["B4"] = 3224
    ws["B5"] = 640
    ws["B6"] = 9
    ws["B7"] = ""

    ws["A9"] = "Impervious area"
    ws["B9"] = "=SUM(B4:B7)"
    ws["C9"] = '=IF(OR(B9="",B9=0),"",B9/43560)'
    ws["D9"] = '=IF(OR(B9="",B9=0,$B$3="",$B$3=0),"",B9/$B$3)'
    for addr in ("A9", "B9", "C9", "D9"):
        ws[addr].font = bold_font
        ws[addr].alignment = right if addr != "A9" else left
    ws["B9"].number_format = "#,##0"
    ws["C9"].number_format = "0.000"
    ws["D9"].number_format = "0.0%"

    ws["A10"] = "Open space"
    ws["B10"] = "=$B$3-$B$9"
    ws["C10"] = '=IF(OR(B10="",B10=0),"",B10/43560)'
    ws["D10"] = '=IF(OR(B10="",B10=0,$B$3="",$B$3=0),"",B10/$B$3)'
    for addr in ("A10", "B10", "C10", "D10"):
        ws[addr].font = bold_font
        ws[addr].alignment = right if addr != "A10" else left
    ws["B10"].number_format = "#,##0"
    ws["C10"].number_format = "0.000"
    ws["D10"].number_format = "0.0%"

    apply_border("A2:D10")

    # -----------------------------
    # STORMWATER RETENTION
    # -----------------------------
    ws["A13"] = "STORMWATER RETENTION"
    ws["A13"].font = title_font
    ws["A13"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[13].height = 22
    ws.merge_cells("A13:C13")

    ws["A24"] = "Retention basis"
    ws["C24"] = "MAX"
    ws["C24"].alignment = left
    ws.merge_cells("C24:D24")

    ws["A25"] = "Side slope ratio (H:V)"
    ws["C25"] = 3
    ws["C25"].alignment = left
    ws.merge_cells("C25:D25")

    ws["A26"] = "Input highlight"
    ws["C26"] = "ON"
    ws["C26"].alignment = left
    ws.merge_cells("C26:D26")

    dv_basis = DataValidation(type="list", formula1='"MAX,LOT ONLY,IMP ONLY"', allow_blank=False)
    ws.add_data_validation(dv_basis)
    dv_basis.add(ws["C24"])

    dv_highlight = DataValidation(type="list", formula1='"ON,OFF"', allow_blank=False)
    ws.add_data_validation(dv_highlight)
    dv_highlight.add(ws["C26"])

    ws["A14"] = "Required (cf)"
    ws["A14"].font = bold_font
    ws["A14"].alignment = left
    ws["B14"] = (
        '=IF($C$24="LOT ONLY",(0.5/12)*$B$3,'
        'IF($C$24="IMP ONLY",(1/12)*$B$9,'
        'MAX((0.5/12)*$B$3,(1/12)*$B$9)))'
    )
    ws["B14"].number_format = "#,##0.0"
    ws["B14"].alignment = right

    ws["A15"] = "Provided (cf)"
    ws["A15"].font = bold_font
    ws["A15"].alignment = left
    # Provided = sum volumes where Select=YES (H19:H22, I19:I22)
    ws["B15"] = '=SUMPRODUCT(--($I$19:$I$22="YES"),$H$19:$H$22)'
    ws["B15"].number_format = "#,##0.0"
    ws["B15"].alignment = right

    ws["A16"] = CellRichText(
        TextBlock(InlineFont(b=True), "FHA Type B:"),
        TextBlock(InlineFont(b=False), " Drainage to front and rear"),
    )
    ws["A16"].alignment = left
    ws.merge_cells("A16:C16")

    apply_border("A14:B15")

    ws["E14"] = (
        '="Required retention uses the larger of two methods:"&CHAR(10)&'
        '"       (a) 1/2"" over total lot area "&TEXT($B$3,"#,##0")&" sf"&CHAR(10)&'
        '"       (b) 1"" over impervious area "&TEXT($B$9,"#,##0")&" sf"'
    )
    ws.merge_cells("E14:I16")
    ws["E14"].alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    # -----------------------------
    # SWALES TABLE
    # -----------------------------
    header_row = 18
    headers = [
        ("Swale", "A"),
        ("Type", "B"),
        ("Bot W (ft)", "C"),
        ("Bot L (ft)", "D"),
        ("Width (ft)", "E"),
        ("Length (ft)", "F"),
        ("Depth (in)", "G"),
        ("Volume (cf)", "H"),
        ("Select", "I"),
    ]

    for text, col in headers:
        cell = f"{col}{header_row}"
        ws[cell] = text
        ws[cell].font = header_font
        ws[cell].alignment = center
        ws[cell].fill = fill_header

    swale_defaults = [
        # Trapezoid/Frustum: top WxL in E/F, depth in G (in), bottom C/D auto-calculated at 3H:1V (using H1)
        dict(name="Swale A", type="Trapezoid", tw=8, tl=48),
        dict(name="Swale B", type="Trapezoid", tw=8, tl=24),
        # V-shape: E=top width, F=length, G=max depth (in) auto-calculated at 3H:1V (using H1)
        dict(name="Swale C", type="V-Shape", tw=8, tl=48),
        dict(name="Swale D", type="V-Shape", tw=8, tl=24),
    ]

    for i, s in enumerate(swale_defaults):
        r = header_row + 1 + i  # 19..22
        ws[f"A{r}"] = s["name"]
        ws[f"A{r}"].alignment = left

        ws[f"B{r}"] = s["type"]
        ws[f"B{r}"].alignment = center

        # inputs
        for col in ("C", "D", "E", "F", "G"):
            ws[f"{col}{r}"].alignment = right

        ws[f"E{r}"] = s["tw"]
        ws[f"F{r}"] = s["tl"]
        if s["type"] == "Trapezoid":
            ws[f"G{r}"] = f'=IF(OR(E{r}="",F{r}=""),"",12*MAX(0,MIN((E{r}-2)/(2*$C$25),(F{r}-2)/(2*$C$25))))'
            ws[f"C{r}"] = f'=IF(OR(E{r}="",G{r}=""),"",MAX(0,E{r}-2*$C$25*(G{r}/12)))'
            ws[f"D{r}"] = f'=IF(OR(F{r}="",G{r}=""),"",MAX(0,F{r}-2*$C$25*(G{r}/12)))'
        else:
            ws[f"G{r}"] = f'=IF(E{r}="","",12*(E{r}/(2*$C$25)))'
        ws[f"C{r}"].number_format = "0.0"
        ws[f"D{r}"].number_format = "0.0"
        ws[f"G{r}"].number_format = "0.0"

        # Volume formula:
        # - V-Shape (with sloped short sides): prismoid along length,
        #   h*TopWidth*(2*TopLength + BottomLength)/6, where BottomLength=max(0, TopLength-TopWidth)
        # - Trapezoid/Frustum: h/3*(A1+A2+sqrt(A1*A2)), A1=C*D, A2=E*F, h=G/12
        ws[f"H{r}"] = (
            f'=IF(UPPER($B{r})="V-SHAPE",'
            f'IF(OR(E{r}="",F{r}="",G{r}=""),"",(G{r}/12)*E{r}*(2*F{r}+MAX(0,F{r}-E{r}))/6),'
            f'IF(OR(C{r}="",D{r}="",E{r}="",F{r}="",G{r}=""),"",'
            f'(G{r}/12)/3*((C{r}*D{r})+(E{r}*F{r})+SQRT((C{r}*D{r})*(E{r}*F{r}))))'
            f')'
        )
        ws[f"H{r}"].number_format = "#,##0.0"
        ws[f"H{r}"].alignment = right

        ws[f"I{r}"] = "NO"
        ws[f"I{r}"].alignment = center

    # YES/NO dropdown
    dv_use = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(dv_use)
    dv_use.add("I19:I22")

    # Guard trapezoid inputs so computed bottom dimensions cannot be below 2.0 ft.
    dv_trap_min_bottom = DataValidation(
        type="custom",
        formula1='OR(UPPER(INDIRECT("B"&ROW()))<>"TRAPEZOID",INDIRECT("E"&ROW())="",INDIRECT("F"&ROW())="",AND(INDIRECT("E"&ROW())>=2,INDIRECT("F"&ROW())>=2))',
        allow_blank=True,
        errorStyle="stop",
    )
    dv_trap_min_bottom.showInputMessage = True
    dv_trap_min_bottom.showErrorMessage = True
    dv_trap_min_bottom.promptTitle = "Trapezoid Minimum Bottom"
    dv_trap_min_bottom.prompt = "For trapezoids, top width and top length must each be at least 2.0 ft so bottom sides stay >= 2.0 ft."
    dv_trap_min_bottom.errorTitle = "Invalid Trapezoid Top Dimensions"
    dv_trap_min_bottom.error = "Increase top width/length to at least 2.0 ft to keep trapezoid bottom width/length at or above 2.0 ft."
    ws.add_data_validation(dv_trap_min_bottom)
    dv_trap_min_bottom.add("E19:F20")

    # Prevent manual edits to bottom dimensions (calculated cells).
    dv_bottom_locked = DataValidation(type="custom", formula1="FALSE", allow_blank=False, errorStyle="stop")
    dv_bottom_locked.showInputMessage = True
    dv_bottom_locked.showErrorMessage = True
    dv_bottom_locked.promptTitle = "Calculated Cell"
    dv_bottom_locked.prompt = "Bottom Width/Length are calculated automatically and cannot be edited."
    dv_bottom_locked.errorTitle = "Bottom Dimensions Locked"
    dv_bottom_locked.error = "Bottom Width and Bottom Length are calculated from Top dimensions and slope. Edit Width/Length instead."
    ws.add_data_validation(dv_bottom_locked)
    dv_bottom_locked.add("C19:D22")

    ws.conditional_formatting.add(
        "B3:B7",
        FormulaRule(formula=['$C$26="ON"'], fill=fill_input, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        "E19:E22",
        FormulaRule(formula=['$C$26="ON"'], fill=fill_input, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        "F19:F22",
        FormulaRule(formula=['$C$26="ON"'], fill=fill_input, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        "A19:H22",
        FormulaRule(formula=['$I19="NO"'], font=deprecate_font, fill=fill_deprecated),
    )

    apply_border("A18:I22")

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate Swale Calculator workbook.")
    p.add_argument("--out", default="swale_calculator.xlsx", help="Output .xlsx path.")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    out_path = make(args.out)
    print(f"Wrote: {out_path}")