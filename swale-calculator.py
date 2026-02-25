#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation


def make(out: Union[str, Path] = "swale_calculator.xlsx") -> Path:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True  # force Excel recalculation on open

    normal_style = wb._named_styles[0]
    normal_font = copy(normal_style.font)
    normal_font.name = "Roboto"
    normal_font.size = 10
    normal_style.font = normal_font

    ws = wb.active
    ws.title = "Swale Calculator"
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.25, footer=0.25)

    # -----------------------------
    # Styles (simple, inline)
    # -----------------------------
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Use opaque ARGB fills (FF......) so Excel shows them
    fill_header = PatternFill("solid", fgColor=Color(theme=9, tint=0.8))
    fill_input = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
    fill_deprecated = PatternFill(fill_type="solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
    deprecate_font = Font(color="FF7F7F7F", strike=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(widths: dict[str, float]) -> None:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def apply_border(cell_range: str) -> None:
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border

    def apply_left_border(cell_range: str) -> None:
        left_border = Border(left=thin)
        for row in ws[cell_range]:
            for cell in row:
                cell.border = left_border

    # -----------------------------
    # Layout
    # -----------------------------
    set_col_width(
        {
            "A": 21,
            "B": 10,
            "C": 10,
            "D": 10,
            "E": 10,
            "F": 10,
            "G": 10,
            "H": 11,
            "I": 10,
            "J": 10,
        }
    )

    # -----------------------------
    # SITE STATISTICS
    # -----------------------------
    ws["A1"] = "SITE STATISTICS"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:C1")

    ws["A2"] = "Description"
    ws["B2"] = "Sq. Feet"
    ws["C2"] = "Acres"
    ws["D2"] = "Percent"
    for addr in ("A2", "B2", "C2", "D2"):
        ws[addr].font = header_font
        ws[addr].alignment = center
        ws[addr].fill = fill_header

    site_rows = [
        ("Lot size", 3),
        ("Building (incl. patios)", 4),
        ("Driveway", 5),
        ("Equipment pads", 6),
        ("Other impervious", 7),
    ]

    for label, r in site_rows:
        ws[f"A{r}"] = label
        ws[f"A{r}"].alignment = left

        ws[f"B{r}"].number_format = "#,##0"
        ws[f"B{r}"].alignment = right

        ws[f"C{r}"] = f'=IF(OR(B{r}="",B{r}=0),"",B{r}/43560)'
        ws[f"C{r}"].number_format = "0.000"
        ws[f"C{r}"].alignment = right

        # Percent: blank for Lot size row, else fraction of lot
        ws[f"D{r}"] = "" if r == 3 else f'=IF(OR(B{r}="",B{r}=0,$B$3="",$B$3=0),"",B{r}/$B$3)'
        ws[f"D{r}"].number_format = "0.0%"
        ws[f"D{r}"].alignment = right

    # Seed example values (edit as needed)
    ws["B3"] = 6741
    ws["B4"] = 3224
    ws["B5"] = 640
    ws["B6"] = 9
    ws["B7"] = ""

    ws["A9"] = "Impervious area"
    ws["B9"] = "=SUM(B4:B7)"
    ws["C9"] = '=IF(OR(B9="",B9=0),"",B9/43560)'
    ws["D9"] = '=IF(OR(B9="",B9=0,$B$3="",$B$3=0),"",B9/$B$3)'
    for addr in ("A9", "B9", "C9", "D9"):
        ws[addr].font = bold_font
        ws[addr].alignment = right if addr != "A9" else left
    ws["B9"].number_format = "#,##0"
    ws["C9"].number_format = "0.000"
    ws["D9"].number_format = "0.0%"

    ws["A10"] = "Open space"
    ws["B10"] = "=$B$3-$B$9"
    ws["C10"] = '=IF(OR(B10="",B10=0),"",B10/43560)'
    ws["D10"] = '=IF(OR(B10="",B10=0,$B$3="",$B$3=0),"",B10/$B$3)'
    for addr in ("A10", "B10", "C10", "D10"):
        ws[addr].font = bold_font
        ws[addr].alignment = right if addr != "A10" else left
    ws["B10"].number_format = "#,##0"
    ws["C10"].number_format = "0.000"
    ws["D10"].number_format = "0.0%"

    apply_border("A2:D10")

    # -----------------------------
    # STORMWATER RETENTION
    # -----------------------------
    ws["A12"] = "STORMWATER RETENTION"
    ws["A12"].font = title_font
    ws["A12"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[12].height = 22
    ws.merge_cells("A12:C12")

    ws["A16"] = "Retention basis"
    ws["A16"].font = bold_font
    ws["B16"] = 'Max (1/2" lot | 1" impervious)'
    ws.merge_cells("B16:D16")

    apply_left_border("E13:E17")

    ws["A17"] = "Side slope ratio (H:V)"
    ws["A17"].font = bold_font
    ws["B17"] = 3
    ws["B17"].alignment = left

    ws["A25"] = "Input highlight"
    ws["A25"].font = bold_font
    ws["B25"] = "ON"

    dv_basis = DataValidation(
        type="list",
        formula1='"Max (1/2"" lot | 1"" impervious),1/2"" over lot,1"" over lot,1.5"" over lot,1"" over impervious"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_basis)
    dv_basis.add(ws["B16"])

    dv_highlight = DataValidation(type="list", formula1='"ON,OFF"', allow_blank=False)
    ws.add_data_validation(dv_highlight)
    dv_highlight.add(ws["B25"])

    ws["A13"] = "Required (cf)"
    ws["A13"].font = bold_font
    ws["B13"] = (
        '=IF($B$16="1/2"" over lot",(0.5/12)*$B$3,'
        'IF($B$16="1"" over lot",(1/12)*$B$3,'
        'IF($B$16="1.5"" over lot",(1.5/12)*$B$3,'
        'IF($B$16="1"" over impervious",(1/12)*$B$9,'
        'MAX((0.5/12)*$B$3,(1/12)*$B$9)))))'
    )
    ws["B13"].number_format = "#,##0.0"
    ws["B13"].alignment = left

    ws["A14"] = "Provided (cf)"
    ws["A14"].font = bold_font
    # Provided = sum volumes where Select=YES (H20:H23, J20:J23)
    ws["B14"] = '=SUMPRODUCT(--($J$20:$J$23="YES"),$H$20:$H$23)'
    ws["B14"].number_format = "#,##0.0"
    ws["B14"].alignment = left

    ws["A15"] = "FHA Type B"
    ws["A15"].font = bold_font
    ws["B15"] = "Drainage to front and rear"
    ws.merge_cells("B15:D15")

    ws["F13"] = "Retention Options (cf)"
    ws["F13"].font = bold_font
    ws["F13"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells("F13:H13")

    ws["F14"] = '1/2" over lot'
    ws["F15"] = '1" over lot'
    ws["F16"] = '1.5" over lot'
    ws["F17"] = '1" over impervious'
    for addr in ("F14", "F15", "F16", "F17"):
        ws[addr].alignment = left
    ws.merge_cells("F14:G14")
    ws.merge_cells("F15:G15")
    ws.merge_cells("F16:G16")
    ws.merge_cells("F17:G17")

    ws["H14"] = '=TEXT((0.5/12)*$B$3,"#,##0.0")'
    ws["H15"] = '=TEXT((1/12)*$B$3,"#,##0.0")'
    ws["H16"] = '=TEXT((1.5/12)*$B$3,"#,##0.0")'
    ws["H17"] = '=TEXT((1/12)*$B$9,"#,##0.0")'
    for addr in ("H14", "H15", "H16", "H17"):
        ws[addr].alignment = right

    # -----------------------------
    # SWALES TABLE
    # -----------------------------
    header_row = 19
    headers = [
        ("Swale", "A"),
        ("Type", "B"),
        ("Bot W (ft)", "C"),
        ("Bot L (ft)", "D"),
        ("Width (ft)", "E"),
        ("Length (ft)", "F"),
        ("Depth (in)", "G"),
        ("Volume (cf)", "H"),
        ("Select", "J"),
    ]

    for text, col in headers:
        cell = f"{col}{header_row}"
        ws[cell] = text
        ws[cell].font = header_font
        ws[cell].alignment = center
        ws[cell].fill = fill_header

    swale_defaults = [
        # Trapezoid/Frustum: top WxL in E/F, depth in G (in), bottom C/D auto-calculated at 3H:1V (using H1)
        dict(name="Swale A", type="Trapezoid", tw=8, tl=48),
        dict(name="Swale B", type="Trapezoid", tw=8, tl=24),
        # V-shape: E=top width, F=length, G=max depth (in) auto-calculated at 3H:1V (using H1)
        dict(name="Swale C", type="V-Shape", tw=8, tl=48),
        dict(name="Swale D", type="V-Shape", tw=8, tl=24),
    ]

    for i, s in enumerate(swale_defaults):
        r = header_row + 1 + i  # 20..23
        ws[f"A{r}"] = s["name"]
        ws[f"A{r}"].alignment = left

        ws[f"B{r}"] = s["type"]
        ws[f"B{r}"].alignment = center

        # inputs
        for col in ("C", "D", "E", "F", "G"):
            ws[f"{col}{r}"].alignment = right

        ws[f"E{r}"] = s["tw"]
        ws[f"F{r}"] = s["tl"]
        if s["type"] == "Trapezoid":
            ws[f"G{r}"] = f'=IF(OR(E{r}="",F{r}=""),"",12*MAX(0,MIN((E{r}-2)/(2*$B$17),(F{r}-2)/(2*$B$17))))'
            ws[f"C{r}"] = f'=IF(OR(E{r}="",G{r}=""),"",MAX(0,E{r}-2*$B$17*(G{r}/12)))'
            ws[f"D{r}"] = f'=IF(OR(F{r}="",G{r}=""),"",MAX(0,F{r}-2*$B$17*(G{r}/12)))'
        else:
            ws[f"G{r}"] = f'=IF(E{r}="","",12*(E{r}/(2*$B$17)))'
        ws[f"C{r}"].number_format = "0.0"
        ws[f"D{r}"].number_format = "0.0"
        ws[f"G{r}"].number_format = "0.0"

        # Volume formula:
        # - V-Shape (with sloped short sides): prismoid along length,
        #   h*TopWidth*(2*TopLength + BottomLength)/6, where BottomLength=max(0, TopLength-TopWidth)
        # - Trapezoid/Frustum: h/3*(A1+A2+sqrt(A1*A2)), A1=C*D, A2=E*F, h=G/12
        ws[f"H{r}"] = (
            f'=IF(UPPER($B{r})="V-SHAPE",'
            f'IF(OR(E{r}="",F{r}="",G{r}=""),"",(G{r}/12)*E{r}*(2*F{r}+MAX(0,F{r}-E{r}))/6),'
            f'IF(OR(C{r}="",D{r}="",E{r}="",F{r}="",G{r}=""),"",'
            f'(G{r}/12)/3*((C{r}*D{r})+(E{r}*F{r})+SQRT((C{r}*D{r})*(E{r}*F{r}))))'
            f')'
        )
        ws[f"H{r}"].number_format = "#,##0.0"
        ws[f"H{r}"].alignment = right

        ws[f"J{r}"] = "NO"
        ws[f"J{r}"].alignment = center

    # YES/NO dropdown
    dv_use = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(dv_use)
    dv_use.add("J20:J23")

    # Guard trapezoid inputs so computed bottom dimensions cannot be below 2.0 ft.
    dv_trap_min_bottom = DataValidation(
        type="custom",
        formula1='OR(UPPER(INDIRECT("B"&ROW()))<>"TRAPEZOID",INDIRECT("E"&ROW())="",INDIRECT("F"&ROW())="",AND(INDIRECT("E"&ROW())>=2,INDIRECT("F"&ROW())>=2))',
        allow_blank=True,
        errorStyle="stop",
    )
    dv_trap_min_bottom.showInputMessage = True
    dv_trap_min_bottom.showErrorMessage = True
    dv_trap_min_bottom.promptTitle = "Trapezoid Minimum Bottom"
    dv_trap_min_bottom.prompt = "For trapezoids, top width and top length must each be at least 2.0 ft so bottom sides stay >= 2.0 ft."
    dv_trap_min_bottom.errorTitle = "Invalid Trapezoid Top Dimensions"
    dv_trap_min_bottom.error = "Increase top width/length to at least 2.0 ft to keep trapezoid bottom width/length at or above 2.0 ft."
    ws.add_data_validation(dv_trap_min_bottom)
    dv_trap_min_bottom.add("E20:F21")

    # Prevent manual edits to bottom dimensions (calculated cells).
    dv_bottom_locked = DataValidation(type="custom", formula1="FALSE", allow_blank=False, errorStyle="stop")
    dv_bottom_locked.showInputMessage = True
    dv_bottom_locked.showErrorMessage = True
    dv_bottom_locked.promptTitle = "Calculated Cell"
    dv_bottom_locked.prompt = "Bottom Width/Length are calculated automatically and cannot be edited."
    dv_bottom_locked.errorTitle = "Bottom Dimensions Locked"
    dv_bottom_locked.error = "Bottom Width and Bottom Length are calculated from Top dimensions and slope. Edit Width/Length instead."
    ws.add_data_validation(dv_bottom_locked)
    dv_bottom_locked.add("C20:D23")

    ws.conditional_formatting.add(
        "B3:B7",
        FormulaRule(formula=['$B$25="ON"'], fill=fill_input, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        "E20:E23",
        FormulaRule(formula=['$B$25="ON"'], fill=fill_input, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        "G20:G23",
        FormulaRule(formula=['$B$25="ON"'], fill=fill_input, stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        "A20:H23",
        FormulaRule(formula=['$J20="NO"'], font=deprecate_font, fill=fill_deprecated),
    )

    apply_border("A19:H23")
    apply_border("J19:J23")

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate Swale Calculator workbook.")
    p.add_argument("--out", default="swale_calculator.xlsx", help="Output .xlsx path.")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    out_path = make(args.out)
    print(f"Wrote: {out_path}")